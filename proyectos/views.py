from .services import obtener_candidatos_optimos
from django.shortcuts import render, get_object_or_404
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST
from django.core.serializers.json import DjangoJSONEncoder
from .models import Tarea, Proyecto
from rrhh.models import Recurso, Perfil, Habilidad, Conocimiento
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import redirect, render
from django.contrib import messages
from django.db.models import Count, Q, Avg
from django.utils import timezone
from datetime import date, datetime
from django.http import HttpResponse
import openpyxl
from django.contrib.auth.decorators import login_required
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .services import procesar_excel_recursos, procesar_excel_proyectos, obtener_kpis_dashboard, obtener_datos_reporte_recursos, generar_excel_reporte
import json

@login_required
@never_cache
def vista_gantt(request):
    # 1. Capturar filtros de la URL
    proyecto_id = request.GET.get('proyecto')
    recurso_id = request.GET.get('recurso')
    
    # 2. Consulta Base
    tareas = Tarea.objects.all().select_related('proyecto', 'asignado_a')
    
    # 3. Aplicar Filtros
    if proyecto_id:
        tareas = tareas.filter(proyecto_id=proyecto_id)
    if recurso_id:
        tareas = tareas.filter(asignado_a_id=recurso_id)
        
    # 4. Preparar datos para JavaScript (Frappe Gantt format)
    gantt_data = []
    for t in tareas:
        gantt_data.append({
            'id': str(t.id),
            'name': f"{t.nombre}",
            'start': t.fecha_inicio.strftime("%Y-%m-%d"),
            'end': t.fecha_fin.strftime("%Y-%m-%d"),
            'progress': t.progreso,
            # Usamos colores diferentes según el proyecto 
            'custom_class': f'bar-project-{t.proyecto.id % 5}', 
            # Información extra para el tooltip
            'proyecto': t.proyecto.nombre,
            'responsable': t.asignado_a.nombre if t.asignado_a else "Sin asignar"
        })
        
    # 5. Listas para los selectores del filtro
    proyectos = Proyecto.objects.all()
    recursos = Recurso.objects.all()

    contexto = {
        'gantt_data': json.dumps(gantt_data, cls=DjangoJSONEncoder),
        'proyectos': proyectos,
        'recursos': recursos,
        'filtros': {'proyecto': proyecto_id, 'recurso': recurso_id} # Para mantener seleccionado el filtro
    }
    
    return render(request, 'proyectos/gantt.html', contexto)

@never_cache
@login_required
def buscar_disponibilidad(request):
    # 1. Captura de Parámetros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    perfil_id = request.GET.get('perfil_id')
    tarea_id = request.GET.get('tarea_id')
    
    tarea_obj = None
    requisitos = []
    mensaje = ""
    candidatos = []

    # 2. Validación de Tarea (Lógica de vista: Redirección)
    if tarea_id:
        try:
            tarea_obj = Tarea.objects.get(id=tarea_id)
            # Candado de seguridad
            if tarea_obj.progreso >= 100:
                messages.warning(request, f"La tarea '{tarea_obj.nombre}' ya finalizó.")
                return redirect('gantt')
            
            # Autocompletar fechas si vienen vacías
            if not fecha_inicio: fecha_inicio = tarea_obj.fecha_inicio.strftime('%Y-%m-%d')
            if not fecha_fin: fecha_fin = tarea_obj.fecha_fin.strftime('%Y-%m-%d')
            
            requisitos = tarea_obj.requisitos.all()
            
        except Tarea.DoesNotExist:
            tarea_obj = None

    # 3. LLAMADA AL SERVICIO (Aquí ocurre la magia)
    if fecha_inicio and fecha_fin:
        try:
            candidatos = obtener_candidatos_optimos(
                fecha_inicio, 
                fecha_fin, 
                perfil_id, 
                requisitos
            )
            if not candidatos:
                mensaje = "No se encontraron recursos que coincidan con los filtros."
        except ValueError as e:
            mensaje = str(e) # "La fecha inicio no puede ser mayor..."

    # 4. Contexto para el HTML
    contexto = {
        'perfiles': Perfil.objects.all(),
        'candidatos': candidatos,
        'mensaje': mensaje,
        'tarea_seleccionada': tarea_obj,
        'filtros': {
            'fecha_inicio': fecha_inicio, 
            'fecha_fin': fecha_fin, 
            'perfil': int(perfil_id) if perfil_id else ''
        }
    }
    
    return render(request, 'proyectos/buscar.html', contexto)

@never_cache
@login_required
@require_POST
def actualizar_tarea_api(request):
    try:
        data = json.loads(request.body)
        tarea_id = data.get('id')
        nueva_fecha_inicio = data.get('start')
        nueva_fecha_fin = data.get('end')
        
        tarea = Tarea.objects.get(id=tarea_id)
        tarea.fecha_inicio = nueva_fecha_inicio[:10]
        tarea.fecha_fin = nueva_fecha_fin[:10]
        tarea.save()
        
        return JsonResponse({'status': 'ok'})
    except Tarea.DoesNotExist:
        return JsonResponse({'status': 'error', 'mensaje': 'Tarea no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)

@never_cache
@login_required
def asignar_recurso(request, tarea_id, recurso_id):
    # Usamos get_object_or_404 por seguridad, aunque .get() también funciona
    tarea = get_object_or_404(Tarea, id=tarea_id)
    recurso = get_object_or_404(Recurso, id=recurso_id)
    
    # 🔒 CANDADO 2: Validación final
    # Evita que alguien fuerce la asignación escribiendo la URL manualmente
    if tarea.progreso >= 100:
        messages.error(request, f"Acción denegada: La tarea '{tarea.nombre}' ya está finalizada.")
        return redirect('gantt')

    tarea.asignado_a = recurso
    tarea.save()

    messages.success(request, f"¡Éxito! La tarea '{tarea.nombre}' ha sido asignada a {recurso.nombre}.")
    
    # Nos devuelve al Gantt para ver el cambio
    return redirect('gantt')

@login_required
@never_cache
def index(request):
    # Delegamos la lógica al servicio
    datos_dashboard = obtener_kpis_dashboard()
    
    # Pasamos los datos al template
    return render(request, 'proyectos/index.html', datos_dashboard)

@login_required
@never_cache
def ver_recursos(request):
    hoy = timezone.now().date()
    recursos = Recurso.objects.all()
    info_recursos = []
    
    for r in recursos:
        # 1. Tareas ACTIVAS (Las que hacen que esté "Ocupado" HOY)
        # Criterio: Ya empezó, no ha terminado su fecha fin, y no está al 100%
        tareas_activas = Tarea.objects.filter(
            asignado_a=r,
            fecha_inicio__lte=hoy,
            fecha_fin__gte=hoy,
            progreso__lt=100
        )

        # 2. Tareas FUTURAS (Solo informativas, no afectan el estado de hoy)
        # Criterio: Empiezan DESPUÉS de hoy
        tareas_futuras = Tarea.objects.filter(
            asignado_a=r,
            fecha_inicio__gt=hoy,
            progreso__lt=100
        ).order_by('fecha_inicio')[:3] 
        
        # Determinar el estado según las tareas ACTIVAS
        estado_actual = 'Ocupado' if tareas_activas.exists() else 'Disponible'

        info_recursos.append({
            'perfil': r,
            'estado': estado_actual,
            'tareas_activas': tareas_activas,
            'tareas_futuras': tareas_futuras 
        })
    
    contexto = {'lista_recursos': info_recursos}
    return render(request, 'proyectos/recursos.html', contexto)

@login_required
def lista_proyectos(request):
    # Traemos proyectos con sus tareas pre-cargadas para optimizar velocidad
    proyectos = Proyecto.objects.prefetch_related('tareas__asignado_a').all()
    
    # Calculamos el avance promedio del proyecto (Opcional, pero útil)
    for p in proyectos:
        total_tareas = p.tareas.count()
        if total_tareas > 0:
            suma_progreso = sum(t.progreso for t in p.tareas.all())
            p.avance_total = round(suma_progreso / total_tareas)
        else:
            p.avance_total = 0

    return render(request, 'proyectos/lista_proyectos.html', {'proyectos': proyectos})

@login_required
@never_cache
def reporte_recurso(request):
    # 1. Capturar filtros
    recurso_id = request.GET.get('recurso')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    es_excel = request.GET.get('exportar') == 'excel'
    
    # 2. Obtener los datos PROCESADOS desde el servicio
    # (Ya no calculamos nada aquí, solo pedimos datos)
    datos_reporte = []
    if recurso_id or request.GET: # Solo buscar si hay filtros o petición explícita
        datos_reporte = obtener_datos_reporte_recursos(recurso_id, fecha_inicio, fecha_fin)

    # 3. ¿Pidieron Excel?
    if es_excel and datos_reporte:
        # El servicio se encarga de crear el archivo y la respuesta HTTP
        return generar_excel_reporte(datos_reporte)

    # 4. ¿Pidieron Web? Renderizamos HTML
    contexto = {
        'recursos': Recurso.objects.all(), # Para llenar el select del filtro
        'lista_reportes': datos_reporte, 
        'mostrar_reporte': len(datos_reporte) > 0,
        'hoy': date.today(),
        'filtros': {
            'recurso': int(recurso_id) if recurso_id else "",
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin
        }
    }

    return render(request, 'proyectos/reporte_recurso.html', contexto)

@login_required
@never_cache
def centro_importacion(request):
    # Esta vista solo renderiza la pantalla con las opciones
    return render(request, 'proyectos/importar_datos.html')

@login_required
@never_cache
def importar_recursos_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        # 1. Llamamos al servicio (él hace el trabajo sucio)
        resultado = procesar_excel_recursos(request.FILES['archivo_excel'])
        
        # 2. La vista solo se encarga de los mensajes
        msg = f"Proceso: {resultado['creados']} creados, {resultado['actualizados']} actualizados."
        if resultado['errores']:
            detalle = " | ".join(resultado['errores'][:2])
            messages.warning(request, f"{msg} Errores: {detalle}")
        else:
            messages.success(request, msg)
            
    return redirect('centro_importacion')

@login_required
@never_cache
def descargar_plantilla_recursos(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Recursos"
    
    # --- ENCABEZADOS ---
    headers = [
        'Nombre Completo', 
        'Cargo/Perfil', 
        'Correo Electrónico', 
        'Habilidades (Formato: Skill:1-5, Skill:1-5)' 
    ]
    ws.append(headers)
    
    # --- EJEMPLO ---
    ws.append([
        'Juan Perez', 
        'Desarrollador Senior', 
        'juan@indutronica.com', 
        'Python:5, Django:4, SQL:3, Liderazgo:5' 
    ])
    
    # Ajustar ancho de columnas para que se lea bien
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 50 

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        content=buffer.getvalue(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Plantilla_Recursos.xlsx'
    return response

@login_required
@never_cache
def importar_proyectos_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        # 1. Llamamos al servicio
        resultado = procesar_excel_proyectos(request.FILES['archivo_excel'])
        
        # 2. Mensajes
        msg = f"Proyectos: {resultado['creados']} creados, {resultado['actualizados']} actualizados."
        if resultado['errores']:
            detalle = " | ".join(resultado['errores'][:2])
            messages.warning(request, f"{msg} Errores: {detalle}")
        else:
            messages.success(request, msg)
            
    return redirect('centro_importacion')

@login_required
def descargar_plantilla_proyectos(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Proyectos"
    
    # Encabezados: Agregamos la Columna E para la Unidad de Negocio
    headers = [
        'Nombre del Proyecto', 
        'Centro de Costo', 
        'Fecha Inicio (YYYY-MM-DD)', 
        'Fecha Fin (YYYY-MM-DD)', 
        'Unidad de Negocio'  # <--- COLUMNA E (NUEVA)
    ]
    ws.append(headers)
    
    # Ejemplos: Incluimos ejemplos válidos para guiar al usuario
    ws.append(['Mantenimiento Norte', 'Obras Civiles', '2026-02-01', '2026-05-30', 'ENERGIA'])
    ws.append(['Instalación CCTV', 'Tecnología', '2026-03-10', '2026-04-10', 'TELECOMUNICACIONES'])
    ws.append(['Tableros PLC', 'Planta 1', '2026-03-15', '2026-06-20', 'AUTOMATIZACION'])
    
    # Ajuste visual de anchos de columna
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25

    # Preparamos el archivo en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        content=buffer.getvalue(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Plantilla_Proyectos_Completa.xlsx'
    return response