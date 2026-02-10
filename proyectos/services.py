# proyectos/services.py
from datetime import datetime, date
from django.db.models import Count, Q
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse
from datetime import datetime
from rrhh.models import Recurso, Habilidad, Conocimiento, Perfil
from .models import Proyecto, Tarea

def procesar_excel_recursos(archivo_excel):
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb.active
    creados = 0
    actualizados = 0
    errores = []

    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            nombre = row[0]
            nombre_perfil = row[1]
            email_excel = row[2]
            habilidades_str = row[3]

            if not nombre: continue

            # Lógica Perfil
            perfil_obj = None
            if nombre_perfil:
                perfil_obj, _ = Perfil.objects.get_or_create(nombre=nombre_perfil)

            # Lógica Recurso
            recurso, created = Recurso.objects.update_or_create(
                nombre=nombre,
                defaults={
                    'perfil': perfil_obj,
                    'email': email_excel if email_excel else "",
                    'activo': True
                }
            )
            if created: creados += 1
            else: actualizados += 1

            # Lógica Habilidades
            if habilidades_str:
                lista_skills = str(habilidades_str).split(',')
                for skill_raw in lista_skills:
                    skill_data = skill_raw.strip()
                    if not skill_data: continue
                    
                    if ':' in skill_data:
                        partes = skill_data.split(':')
                        nombre_skill = partes[0].strip()
                        try: nivel_skill = int(partes[1].strip())
                        except: nivel_skill = 1
                    else:
                        nombre_skill = skill_data
                        nivel_skill = 1

                    conocimiento_obj, _ = Conocimiento.objects.get_or_create(nombre=nombre_skill)
                    Habilidad.objects.update_or_create(recurso=recurso, conocimiento=conocimiento_obj, defaults={'nivel': nivel_skill})

        except Exception as e:
            errores.append(f"Fila {index} ({nombre}): {str(e)}")

    return {'creados': creados, 'actualizados': actualizados, 'errores': errores}

def procesar_excel_proyectos(archivo_excel):
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb.active
    creados = 0
    actualizados = 0
    errores = []

    def limpiar_fecha(valor, default):
        if not valor: return default
        if isinstance(valor, datetime) or hasattr(valor, 'date'): return valor
        try: return datetime.strptime(str(valor).strip(), '%Y-%m-%d').date()
        except: return default

    hoy = datetime.now().date()

    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            nombre = row[0]
            centro_costo = row[1]
            f_inicio_raw = row[2]
            f_fin_raw = row[3]
            unidad_raw = row[4]

            if not nombre: continue

            f_inicio = limpiar_fecha(f_inicio_raw, hoy)
            f_fin = limpiar_fecha(f_fin_raw, hoy)

            unidad_final = 'AUTOMATIZACION'
            if unidad_raw:
                texto_limpio = str(unidad_raw).upper().strip()
                mapa_unidades = {
                    'ENERGIA': 'ENERGIA', 'ENERGÍA': 'ENERGIA',
                    'TELECOMUNICACIONES': 'TELECOMUNICACIONES', 'TELECOM': 'TELECOMUNICACIONES',
                    'AUTOMATIZACION': 'AUTOMATIZACION', 'AUTOMATIZACIÓN': 'AUTOMATIZACION'
                }
                unidad_final = mapa_unidades.get(texto_limpio, 'AUTOMATIZACION')

            obj, created = Proyecto.objects.update_or_create(
                nombre=nombre,
                defaults={
                    'centro_costo': centro_costo if centro_costo else "General",
                    'fecha_inicio': f_inicio,
                    'fecha_fin_estimada': f_fin,
                    'unidad_negocio': unidad_final,
                }
            )
            if created: creados += 1
            else: actualizados += 1

        except Exception as e:
            errores.append(f"Fila {index} ({nombre}): {str(e)}")

    return {'creados': creados, 'actualizados': actualizados, 'errores': errores}

def obtener_kpis_dashboard():
    """
    Calcula KPIs, Gráficas y Lista de Proyectos agrupados para el dashboard.
    """
    # 1. KPIs Generales
    total_proyectos = Proyecto.objects.count()
    total_recursos = Recurso.objects.filter(activo=True).count()
    tareas_pendientes = Tarea.objects.filter(progreso__lt=100).count()
    tareas_completadas = Tarea.objects.filter(progreso=100).count() # Recuperamos este dato
    total_tareas = Tarea.objects.count() # Recuperamos este dato

    # 2. Datos para Gráficas (Dona y Barras)
    data_proyectos = Proyecto.objects.values('unidad_negocio').annotate(total=Count('id'))
    labels_proyectos = [item['unidad_negocio'] for item in data_proyectos]
    values_proyectos = [item['total'] for item in data_proyectos]

    recursos_ocupados = Recurso.objects.annotate(
        num_tareas=Count('tarea', filter=Q(tarea__progreso__lt=100))
    ).order_by('-num_tareas')[:5]
    labels_recursos = [r.nombre.split()[0] for r in recursos_ocupados]
    values_recursos = [r.num_tareas for r in recursos_ocupados]

    # 3. Lógica del Acordeón (Proyectos por Centro de Costo)
    # Recuperamos la lógica que tenías antes
    proyectos_por_cc = []
    centros_unicos = Proyecto.objects.values_list('centro_costo', flat=True).distinct()
    
    for cc in centros_unicos:
        if cc: # Evitar vacíos
            projs = Proyecto.objects.filter(centro_costo=cc).order_by('fecha_fin_estimada')
            if projs.exists():
                proyectos_por_cc.append({
                    'nombre_cc': cc,
                    'cantidad': projs.count(),
                    'lista_proyectos': projs
                })

    # Retornamos TODO en un solo diccionario ordenado
    return {
        'kpis': {
            'total_proyectos': total_proyectos,
            'total_recursos': total_recursos,
            'total_tareas': total_tareas,
            'tareas_pendientes': tareas_pendientes,
            'tareas_completadas': tareas_completadas,
        },
        'graficas': {
            'labels_proyectos': labels_proyectos,
            'values_proyectos': values_proyectos,
            'labels_recursos': labels_recursos,
            'values_recursos': values_recursos,
        },
        'proyectos_por_cc': proyectos_por_cc # Aquí va tu lista antigua
    }

def obtener_candidatos_optimos(fecha_inicio_str, fecha_fin_str, perfil_id=None, requisitos=None):
    """
    Motor de búsqueda de recursos:
    1. Filtra por perfil.
    2. Detecta ocupación.
    3. Calcula Match Técnico (Skills).
    4. Ordena por Disponibilidad > Puntaje.
    """
    candidatos_finales = []
    
    # 1. Validación de fechas
    if not fecha_inicio_str or not fecha_fin_str:
        return []
    
    if fecha_inicio_str > fecha_fin_str:
        raise ValueError("La fecha de inicio no puede ser mayor al fin.")

    # 2. Buscar universo de recursos activos
    recursos = Recurso.objects.filter(activo=True)
    if perfil_id:
        recursos = recursos.filter(perfil_id=perfil_id)

    # 3. Identificar IDs ocupados en ese rango (Query optimizada)
    ocupados_ids = Tarea.objects.filter(
        progreso__lt=100,
        fecha_inicio__lte=fecha_fin_str,
        fecha_fin__gte=fecha_inicio_str
    ).values_list('asignado_a_id', flat=True)

    # 4. Procesar cada recurso (Algoritmo de Match)
    for recurso in recursos:
        esta_ocupado = recurso.id in ocupados_ids
        
        # A. Datos de Ocupación (Helper function abajo)
        info_ocupacion = _obtener_info_liberacion(recurso, esta_ocupado, fecha_inicio_str)
        
        # B. Datos de Skills (Helper function abajo)
        info_skills = _calcular_match_tecnico(recurso, requisitos)

        # C. Construir candidato
        candidatos_finales.append({
            'perfil': recurso,
            'ocupado': esta_ocupado,
            'fecha_liberacion': info_ocupacion['fecha'],
            'tarea_actual': info_ocupacion['tarea'],
            'match': info_skills['score'],
            'detalles': info_skills['detalles']
        })

    # 5. Ordenamiento Inteligente: Primero los Libres, luego por mejor Match
    candidatos_finales.sort(key=lambda x: (not x['ocupado'], x['match']), reverse=True)
    
    return candidatos_finales

# --- FUNCIONES PRIVADAS (HELPERS) ---

def _obtener_info_liberacion(recurso, esta_ocupado, fecha_inicio_limite):
    """Calcula cuándo se libera un recurso ocupado."""
    if not esta_ocupado:
        return {'fecha': None, 'tarea': ''}
    
    ultima_tarea = Tarea.objects.filter(
        asignado_a=recurso,
        progreso__lt=100,
        fecha_fin__gte=fecha_inicio_limite
    ).order_by('-fecha_fin').first()
    
    if ultima_tarea:
        return {'fecha': ultima_tarea.fecha_fin, 'tarea': ultima_tarea.nombre}
    return {'fecha': None, 'tarea': ''}

def _calcular_match_tecnico(recurso, requisitos):
    """Calcula el porcentaje de coincidencia de habilidades."""
    if not requisitos:
        return {'score': 100, 'detalles': []} # Si no piden nada, cumple al 100%

    puntos_totales = 0
    detalles = []
    
    for req in requisitos:
        # Buscamos si el recurso tiene esa habilidad específica
        habilidad = Habilidad.objects.filter(recurso=recurso, conocimiento=req).first()
        
        if habilidad:
            # Asumiendo que nivel va de 1 a 5
            puntos = (habilidad.nivel / 5) * 100
            puntos_totales += puntos
            detalles.append({
                'skill': req.nombre, 
                'nivel': habilidad.get_nivel_display(), 
                'cumple': True
            })
        else:
            detalles.append({
                'skill': req.nombre, 
                'nivel': '---', 
                'cumple': False
            })
            
    avg_score = round(puntos_totales / len(requisitos))
    return {'score': avg_score, 'detalles': detalles}

def obtener_datos_reporte_recursos(recurso_id=None, fecha_inicio=None, fecha_fin=None):
    """
    Busca recursos y calcula sus estadísticas de tareas según los filtros.
    Retorna una lista de diccionarios con la estructura lista para usar.
    """
    # 1. Filtro inicial de recursos
    if recurso_id:
        recursos_filtrados = Recurso.objects.filter(id=recurso_id)
    else:
        # Si no hay ID, traemos todos (siempre que la función sea llamada con intención de reporte)
        recursos_filtrados = Recurso.objects.all()

    datos_reporte = []

    # 2. Procesamiento
    for recurso in recursos_filtrados:
        tareas = Tarea.objects.filter(asignado_a=recurso).order_by('-fecha_fin')
        
        if fecha_inicio:
            tareas = tareas.filter(fecha_inicio__gte=fecha_inicio)
        if fecha_fin:
            tareas = tareas.filter(fecha_fin__lte=fecha_fin)
            
        # Estadísticas
        total_tareas = tareas.count()
        if total_tareas == 0 and recurso_id is None:
            continue # Opcional: Si es reporte general, saltar los que no tienen nada

        completadas = tareas.filter(progreso=100).count()
        pendientes = total_tareas - completadas
        rendimiento = round((completadas / total_tareas * 100), 1) if total_tareas > 0 else 0
        
        datos_reporte.append({
            'recurso': recurso,
            'tareas': tareas,
            'stats': {
                'total': total_tareas,
                'completadas': completadas,
                'pendientes': pendientes,
                'rendimiento': rendimiento
            }
        })
    
    return datos_reporte

def generar_excel_reporte(datos_reporte):
    """
    Recibe la lista de datos procesados y devuelve una respuesta HTTP con el archivo Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Recursos"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    
    # Encabezados
    headers = ['Ingeniero/Recurso', 'Cargo/Perfil', 'Proyecto', 'Tarea', 'Inicio', 'Fin', 'Estado', 'Progreso (%)']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Llenar datos
    hoy = date.today()
    for data in datos_reporte:
        r_nombre = data['recurso'].nombre
        
        # Lógica defensiva para obtener el cargo
        try:
            r_cargo = getattr(data['recurso'], 'cargo', None)
            if not r_cargo:
                r_cargo = str(getattr(data['recurso'], 'perfil', "No especificado"))
        except Exception:
            r_cargo = "No especificado"

        if not data['tareas']:
            ws.append([r_nombre, r_cargo, "Sin tareas", "-", "-", "-", "-", "-"])
            continue

        for t in data['tareas']:
            estado = "En Curso"
            if t.progreso == 100: estado = "Finalizado"
            elif t.fecha_fin < hoy: estado = "Atrasado"

            ws.append([
                r_nombre, r_cargo, t.proyecto.nombre, t.nombre,
                t.fecha_inicio, t.fecha_fin, estado, t.progreso
            ])

    # Ajuste visual
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30

    # Preparar respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Recursos_RMS.xlsx'
    wb.save(response)
    
    return response